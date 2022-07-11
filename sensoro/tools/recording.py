#!/usr/bin/env/python3
# -*- coding:utf-8 -*-

from openpyxl import load_workbook
import os


# Excel中写入数据
class Recording:
    def __init__(self, file):
        self.wb = load_workbook(file)

    def write_excel(self, sheet_name: str, row: int = None, column: int = None, value=None):
        sheet = self.wb.get_sheet_by_name(sheet_name)
        sheet.cell(row=row, column=column).value = value

    # 读取一列内容
    def read_excel_for_column(self, sheet_name, column: str = None):
        sheet = self.wb[sheet_name]

        columns = sheet[column]
        value = []
        for i in columns:
            value.append(i.value)
        return value

    # 读取一行内容
    def read_excel_for_row(self, sheet_name, row: int = None):
        sheet = self.wb[sheet_name]
        rows = sheet[row]
        value = []
        for i in rows:
            value.append(i.value)
        return value

    # 读取部分行 部分列内容,或 某一单元格内容
    def read_excel_for_row_and_column(self, sheet_name, min_row, max_row, min_column, max_column):
        sheet = self.wb[sheet_name]
        for row in sheet.iter_rows(min_row=min_row, max_row=max_row, min_col=min_column, max_col=max_column):
            for cell in row:
                print(cell.value)

    # 读取某一单元格内容
    def read_excel_cell(self, sheet_name, row, column):
        sheet = self.wb[sheet_name]
        return sheet.cell(row=row, column=column).value

    def save(self, file):
        self.wb.save(file)


if __name__ == '__main__':
    print(Recording('contrast_data.xlsx').read_excel_cell('人脸比对数据分析', 3, 2))
