# coding: utf-8
"""
适用范围：
    1、百度云端算法
测试内容：
    1、对象检测数量
    2、对象关联方式
使用环境：
    1、点军环境，因测试环境接口调用需要收费，点军环境百度部署了一套，调用不收费
"""
import time
import datetime
import os
import json
import hmac
import hashlib
import yaml
from openpyxl import load_workbook
from hashlib import sha256

import requests
from PIL import Image

import pymysql
from typing import Union

import base64
import cv2


# 获取api auth
class BaiAuth(object):
    def __init__(self, user=None, password=None, minute_delta=5 * 60 * 1000):
        self.user = user
        self.password = password

        now = time.time()
        expire = int(round(now * 1000)) + minute_delta * 60 * 1000
        self.expire_time = str(expire)
        self.headers = {}

    @staticmethod
    def to_md5(_str):
        md5 = hashlib.md5()
        md5.update(_str.encode("utf-8"))
        return md5.hexdigest()

    @staticmethod
    def to_json(json_object):
        return json.dumps(json_object, separators=(",", ":"), default=str)

    def generate_auth_headers(self, uri, body):
        """ generate auth params
        :param uri:
        :param body:
        :return:
        """

        node_body_str = self.to_json(body)
        node_body = self.to_md5(node_body_str)

        self.headers = {
            "x-evs-user": self.user,
            "x-evs-expire-date": str(self.expire_time),
            "x-evs-content-md5": node_body,
            "Authorization": None,
            "content-type": "application/json;charset=UTF-8",
        }
        content = f"{self.user}:{self.password}:{uri}:{self.expire_time}:{node_body}"
        self.headers["Authorization"] = hmac.new(
            key=self.password.encode("utf-8"),
            msg=content.encode("utf-8"),
            digestmod=sha256
        ).hexdigest()
        return self.headers


class BaiduAi:
    def __init__(self, user_id, user_sk, server_addr):

        self.user_id = user_id
        self.user_sk = user_sk
        self.server_addr = server_addr

    # 获取接口auth
    def get_or_create_request_headers(self, url_path, body, minute_delta=5 * 60):
        return BaiAuth(
            user=self.user_id, password=self.user_sk, minute_delta=minute_delta
        ).generate_auth_headers(uri=url_path, body=body)

    # 接口调用
    def _base_request(self, url_path: str, **kwargs):
        """

        :param url_path: 百度api地址
        :return:
        """
        api = self.server_addr + url_path

        headers = self.get_or_create_request_headers(url_path, kwargs)

        # a = datetime.datetime.now()
        r = requests.post(api, json=kwargs, headers=headers)
        # b = datetime.datetime.now()
        # print((b-a).seconds)
        rep = r.json()

        return rep

    # 判断使用文件方式
    def base_detect(self, url_path: str, image: str, image_type, **kwargs):
        """

        :param url_path: 百度api地址
        :param image: 待检测待图片，可以是url，文件系统路径、base64编码后的字符串3中类型
        :param image_type: 图片参数的类型，["filepath", "url", "base64str"]
        :return:
        """
        if image_type == "base64str":
            # 适用于api处理前端页面上传图片b64
            # 需要调用者保证图片格式是jpg
            params = {"image_base64": image}
        elif image_type == "filepath":
            # 适用于处理本地文件系统中的图片
            # png --> jpg
            with open(image, 'rb') as image_file:
                params = {"image_base64": base64.b64encode(image_file.read()).decode()}

        elif image_type == "url":
            params = {"image_url": image}

        params.update(kwargs)
        return self._base_request(url_path=url_path, **params)


# 一些公共方法
class Util:

    # 文件插入数据库，并存储图片文件分辨率及大小
    def file_insert_db(self, path, image_type, db):

        file_list = []
        for f in db.execute_sql(f"""select image from algorithm where alg_type='{image_type}'"""):
            file_list.append(f[0])

        val = ''    # 保存出入数据
        for root, dirs, files in os.walk(path):

            new_files = [i for i in files if root+i not in file_list]
            # print(new_files)
            new_files.sort()
            if '结果' in root:
                continue
            else:
                num = False  # 用来判断插入是否有数据，无则不执行sql

                if '.DS_Store' in new_files:
                    new_files.remove('.DS_Store')   # 删除Mac自带的文件
                for n in range(len(new_files)):
                    file_path = root + new_files[n]
                    if new_files[n].split('.')[1].lower() in ('png', 'jpg', 'jpeg'):
                        num = True
                        resolution_size = self.get_resolution_size_for_db(file_path)
                        if n < len(new_files)-1:
                            val += f"('{file_path}', '{image_type}', '{resolution_size[0]}', '{resolution_size[1]}'),"
                        else:
                            val += f"('{file_path}', '{image_type}', '{resolution_size[0]}', '{resolution_size[1]}')"

        if num:

            db.execute_sql(f"""insert into algorithm(image, alg_type, resolution, size) values {val}""")
            print("数据保存数据库完成！")
        else:
            print("无新增数据，开始调用接口识别数据！")

    @staticmethod
    def get_resolution_size_for_db(path):
        img = Image.open(path)
        resolution = img.size  # 获取文件分辨率

        # 获取文件大小
        try:
            size = os.path.getsize(path)

            # 计算大小
            byte = float(size)
            kb = byte / 1024
            if kb >= 1024:
                m = kb / 1024
                return f"{resolution[0]}*{resolution[1]}", "%.1fm" % m
            else:
                return f"{resolution[0]}*{resolution[1]}", f"{int(kb)}kb"

        except Exception as e:
            raise e

    # 测试集数据写入Excel
    def test_data_insert_excel(self, excel_path, sheet_name, file_path):
        wb = load_workbook(excel_path)
        sheet = wb[sheet_name]
        row = input("请输入Excel开始插入位置：")
        self.file_list(file_path=file_path, whether_excel=1, row=row, sheet=sheet)
        wb.save(self.read_config()["data_log"])

    # 获取文件大小及分辨率 写入Excel
    @staticmethod
    def get_resolution_size(picture, row, sheet):

        img = Image.open(picture)
        resolution = img.size     # 获取文件分辨率

        sheet.cell(row, 6).value = f"{resolution[0]}*{resolution[1]}"

        # 获取文件大小
        try:
            size = os.path.getsize(picture)

            # 计算大小
            byte = float(size)
            kb = byte/1024
            if kb >= 1024:
                m = kb/1024
                sheet.cell(row, 5).value = "%.1fm" % m
            else:
                sheet.cell(row, 5).value = f"{int(kb)}kb"
        except Exception as e:
            raise e

    @staticmethod
    def update_name(old_file, new_name):
        path = os.path.split(old_file)[0]
        os.rename(old_file, f"{path}/{new_name}")

    # # 获取目录下所有符合要求的图片，根据关键字判断是否将名称写入Excel中
    # def file_list(self, file_path, relative_path='', whether_excel=0, row=1, sheet=None):
    #
    #     all_file = list(os.listdir(f"{file_path}"))
    #
    #     for f in all_file:
    #         if "结果" not in f and f != "货船客船":
    #             path = f"{file_path}{f}"
    #
    #             if os.path.isfile(path):
    #                 if f.split('.')[1].lower() in ('png', 'jpg', 'jpeg'):
    #                     if whether_excel == 1:
    #                         sheet.cell(row, column=1).value = f"{relative_path}{f}"
    #                         row += 1
    #                     else:
    #                         file.append(path)
    #             elif os.path.isdir(path):
    #                 if whether_excel == 1:
    #                     self.file_list(file_path=f"{path}/", relative_path=f"{f}/",
    #                                    whether_excel=1, row=row, sheet=sheet)
    #                 else:
    #                     self.file_list(file_path=f"{path}/")

    # 获取对象坐标

    @staticmethod
    def read_config():
        # 读取配置文件
        with open(f"{os.getcwd()}/config.yml", "r", encoding="utf-8") as y:
            config = yaml.load(y.read(), Loader=yaml.FullLoader)
        return config


class DB:
    mysql = Util.read_config()['database']

    def __init__(self):
        """
        初始化数据库连接，并指定查询结果以字典形式返回
        """
        self.connection = pymysql.connect(
            host=self.mysql['host'],
            port=self.mysql['port'],
            user=self.mysql['user'],
            password=self.mysql['password'],
            db=self.mysql['db_name'],
            charset=self.mysql['charset'],
        )

    def execute_sql(self, sql: str) -> Union[dict, None]:
        """
        执行sql语句方法，查询所有结果的sql只会返回一条结果支持select， delete， insert， update
        :param sql: sql语句
        :return: select语句，如果有结果则返回 对应结果字典，delete，insert，update 将返回None
        """
        with self.connection.cursor() as cursor:
            cursor.execute(sql)
            result = cursor.fetchall()
            # 使用commit解决查询数据概率查错问题
            self.connection.commit()
            return self.verify(result)

    @staticmethod
    def verify(result: dict) -> Union[dict, None]:
        """验证结果能否被json.dumps序列化问题"""
        # 尝试变成字符串，解决datetime无法被json序列化问题
        try:
            json.dumps(result)
        except TypeError:   # TypeError: Object of type datetime is not JSON serializable
            for k, v in result.items():
                if isinstance(v, datetime):
                    result[k] = str[v]
        return result

    def close(self):
        # 关闭数据库连接
        self.connection.close()


class Target:

    def __init__(self, alg_type):
        self.alg_type = alg_type
        self.db = DB()
        self.config = Util.read_config()
        self.item_count = {}    # 记录图片对象数量

    def target(self, is_item_count=0, is_relation_map=0):
        """
        仅全目标使用该参数
        :param is_item_count: 是否显示对象数量
        :param is_relation_map: 是否显示对象关联关系
        :return:
        """

        # 将数据存入数据库，集成到前端平台时，此段代码可去除
        Util().file_insert_db(self.config['file_path'][self.alg_type], self.alg_type, self.db)

        # 调用百度算法接口并将结果保存至数据库
        self._update_response()

        # 查询所有本次算法需要的数据, relation_map在全目标时使用
        images = self.db.execute_sql(
            f'''select id, image, relation_map from algorithm where alg_type='{self.alg_type}' and status=1''')

        # 循环处理数据：对图片标记，并将标记后的图片地址保存至数据库
        for im_id, image, relation_map in images:
            point = {}  # 记录坐标
            # 查询图片中所有对象及对应属性内容，attribute仅全目标使用,class_id、name其他算子使用
            items = self.db.execute_sql(
                f"""select type, type_id, attribute, class_id, name, location, score 
                from algorithm_item where algorithm_id={im_id}""")

            # 获取对象标记坐标
            for item_type, type_id, attribute, class_id, name, location, score in items:
                if self.alg_type == 'full':
                    if item_type == 'human':
                        point['h' + str(type_id)] = self._get_point(json.loads(location))

                    elif item_type == 'face':
                        point['f' + str(type_id)] = self._get_point(json.loads(location))

                    elif item_type == 'car':
                        point['c' + str(type_id)] = self._get_point(json.loads(location))

                    elif item_type == 'electric-car':
                        point['e' + str(type_id)] = self._get_point(json.loads(location))

                elif self.alg_type == 'boat':
                    point['b' + str(type_id)] = self._get_point(json.loads(location))

                elif self.alg_type == 'fish':
                    point['f' + str(type_id)] = self._get_point(json.loads(location))

                elif self.alg_type == 'trash':
                    point['t' + str(type_id)] = self._get_point(json.loads(location))

            # 对图片进行标记，并将新图片位置保存至数据库
            new_image = self._draw_rectangle(image, point)  # 对图像标记并将新文件位置保存至数据库
            self.db.execute_sql(f'''update algorithm set new_image='{new_image}' where id={im_id}''')

            if is_relation_map:
                self._get_relation_map(relation_map)

        self.db.close()

    # 获取对象坐标
    @staticmethod
    def _get_point(location):
        return (
            (int(location['left']), int(location['top'])),
            (
                int(location['left'] + location['width']),
                int(location['top'] + location['height'])
            )
        )

    # 对图片进行标记，并返回标记后图片的保存位置
    @staticmethod
    def _draw_rectangle(image, point):
        img = cv2.imread(image)

        for key, value in point.items():

            if key[0] == 'b':
                color = (255, 255, 0)

            elif key[0] == 'h':
                color = (0, 255, 0)

            elif key[0] == 'f':
                color = (0, 0, 255)

            elif key[0] == 'c':
                color = (255, 0, 0)

            elif key[0] == 'e':
                color = (0, 255, 255)

            elif key[0] == 't':
                color = (0, 255, 255)

            cv2.rectangle(img, value[0], value[1], color, 2, 4)
            cv2.putText(img, key, (value[0][0], value[0][1]+30), cv2.FONT_HERSHEY_COMPLEX, fontScale=1, color=color, thickness=2)

        new_img_file_path = os.path.split(image)[0] + f"/结果/"
        if not os.path.exists(new_img_file_path):
            os.makedirs(new_img_file_path)

        cv2.imwrite(f"{new_img_file_path}{os.path.split(image)[1]}", img)
        return f"{new_img_file_path}{os.path.split(image)[1]}"

    # 将百度接口返回数据保存至数据库
    def _update_response(self):

        baidu = BaiduAi(self.config["server"]["DianJun"]["user"],
                        self.config["server"]["DianJun"]["password"],
                        self.config["server"]["DianJun"]["url"])

        for num in range(5):  # 控制循环5次，防止网络问题导致的失败
            # 查询数据库中未处理过的数据
            result = self.db.execute_sql(
                f'''select id, image from algorithm where alg_type='{self.alg_type}' and status=0''')

            # 判断是否有未处理过的图片，无直接结束循环
            if not result:
                break
            else:
                for im_id, image in result:
                    """
                    百度接口相关参数：
                    min_score	可选	Double	最小置信度，如果指定该参数，会对检测、识别结果按min_score过滤，只保留置信度大于min_score的特征结果
                    
                    以下为全目标可选参数，其他算法接口传入不影响接口调用：
                    enable_face	可选	Boolean	全目标检测是否检测人脸，默认true；如果不关注人脸目标，可以关闭人脸检测
                    enable_human	可选	Boolean	全目标检测是否检测人体，默认true；如果不关注人体目标，可以关闭人体检测
                    enable_car	可选	Boolean	全目标检测是否检测车辆(包括二/三轮车)，默认true；如果不关注车辆目标，可以关闭车辆检测
                    enable_multiple	可选	Boolean	指定大图模式，全目标检测输出所有目标，默认false，即小图模式
                    control_flag	可选	String	对于小图模式，指定最大目标输出策略，可选值：FACE、HUMAN、CAR，即以最大人脸、最大人体，还是最大车辆优先输出
                    """

                    try:
                        response = baidu.base_detect(
                            self.config['url_map'][self.alg_type], image, "filepath", enable_multiple=True)
                    except Exception as e:
                        # raise e
                        print(e, "服务异常，重新执行！")
                        continue
                    # print(response)
                    try:
                        if response["code"] == 0:  # 判断是否成功调用接口
                            if 'data' in response:  # 部分接口未识别到对象时返回无data字段
                                if response["data"]["item_count"] > 0:  # 判断接口是否识别到对象
                                    value = ""  # 参数组装
                                    items = response["data"]["items"]
                                    if self.alg_type == 'full':
                                        for item in items:
                                            # face属性未单独在一个dict中，需要对其进行组装
                                            if item['type'] == 'face':
                                                attribute = self.face_attribute(item)
                                            elif item['type'] == 'electric-car':  # 删除部分非机动车属性
                                                item['attribute'].pop('plate')
                                                item['attribute'].pop('vehicle_type')
                                                attribute = item['attribute']
                                            else:
                                                attribute = item['attribute']

                                            value += f"({im_id}, " \
                                                     f"'{item['type']}'," \
                                                     f" {item['id']}, " \
                                                     f"'{json.dumps(attribute, ensure_ascii=False)}', " \
                                                     f"Null," \
                                                     f"Null," \
                                                     f"'{json.dumps(item['location'])}', " \
                                                     f"{item['score']}),"
                                        relation = {'relation_map': response['data']['relation_map']}
                                        self.db.execute_sql(
                                            f"""update algorithm 
                                            set relation_map='{json.dumps(relation)}' 
                                            where id={im_id}""")

                                    elif self.alg_type in ('boat', 'fish', 'trash'):
                                        for n in range(len(items)):
                                            value += f"({im_id}," \
                                                     f"'{self.alg_type}'," \
                                                     f"{n}," \
                                                     f"Null," \
                                                     f"{items[n]['class_id']}," \
                                                     f"'{items[n]['name']}'," \
                                                     f"'{json.dumps(items[n]['location'])}'," \
                                                     f"{items[n]['score']}),"

                                    self.db.execute_sql(
                                            f"""insert into algorithm_item
                                            (algorithm_id, type, type_id, attribute,class_id, name, location, score) 
                                            values {value[:-1]}""")

                                    self.db.execute_sql(
                                        f"""update algorithm set status=1 where id={im_id}""")

                                else:
                                    self.db.execute_sql(
                                        f'''update algorithm set status=2 where id={im_id}''')
                            else:
                                self.db.execute_sql(
                                    f'''update algorithm set status=2 where id={im_id}''')
                        elif response['code'] != 0:
                            self.db.execute_sql(
                                f"""update algorithm 
                                set baidu_response='{json.dumps(response)}', status=3 where id={im_id}""")

                    except Exception as e:
                        # raise e
                        print(e, "服务异常，重新执行！")
                        continue

    # 人脸属性未保存在attribute，单独处理
    @staticmethod
    def face_attribute(item):

        attribute = {
            '年龄': item['age'],
            '表情': item['expression'],   # none:不笑；smile:微笑；laugh:大笑
            '性别': item['gender'],   # male:男性 female:女性
            '是否戴眼镜': item['glasses'],    # 是否带眼镜, none:无眼镜，common:普通眼镜，sun:墨镜
            '人种': item['race'],     # yellow: 黄种人 white: 白种人 black: 黑种人 arabs: 阿拉伯人
            '口罩': item['mask'],     # 0代表没戴口罩 1代表戴口罩
            '情绪': item['emotion']  # angry:愤怒 disgust:厌恶 fear:恐惧 happy:高兴 sad:伤心 surprise:惊讶 neutral:无表情 pouty: 撅嘴 grimace:鬼脸
        }

        return attribute

    # 对象关联关系处理
    @staticmethod
    def _get_relation_map(relation_map):
        print('对象关联关系：')
        if not relation_map:
            print('对象无关联关系')
        else:
            for relation in relation_map:
                if relation['car_id'] != -1 and relation['human_id'] != -1 and relation['face_id'] != -1:
                    print(f"\t车辆{relation['car_id']} 人体{relation['human_id']} 人脸{relation['face_id']}")

                elif relation['car_id'] != -1 and relation['human_id'] != -1 and relation['face_id'] == -1:
                    print(f"\t车辆{relation['car_id']} 人体{relation['human_id']}")

                elif relation['car_id'] == -1 and relation['human_id'] != -1 and relation['face_id'] != -1:
                    print(f"\t人体{relation['human_id']} 人脸{relation['face_id']}")

                elif relation['car_id'] != -1 and relation['human_id'] == -1 and relation['face_id'] != -1:
                    print(f"\t车辆{relation['car_id']} 人脸{relation['face_id']}")


if __name__ == '__main__':

    print('file processing started！', time.strftime('%Y-%m-%d %H:%M:%S'))

    while True:
        api = input("请选择调用的api\n1、全目标\n2、船只识别：\n3、非法捕鱼\n4、垃圾桶检测\n")
        if api not in ("1", "2", "3", "4"):
            print("选择错误！")
            continue
        else:
            break

    if int(api) == 1:
        Target('full').target()
    elif int(api) == 2:
        Target('boat').target()
    elif int(api) == 3:
        Target("fish").target()
    elif int(api) == 4:
        Target('trash').target()

    print('file processing end！', time.strftime('%Y-%m-%d %H:%M:%S'))

    # Run().run()

    # BaiduAi._base_detect(url, image, 'url', min_score=0.6)


